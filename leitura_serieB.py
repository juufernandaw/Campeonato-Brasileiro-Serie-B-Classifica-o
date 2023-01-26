# import module
import openpyxl

# load excel with its path
wrkbk = openpyxl.load_workbook("tabela_serieB.xlsx")

sh = wrkbk.active
# for sheet in wb.worksheets:
# iterate through excel and display data
for i in range(2, sh.max_row + 1):

    for j in range(1, sh.max_column + 1):
        cell_obj = sh.cell(row=i, column=j)
        rodada = sh.cell(row=i, column=2).value
        mandante = sh.cell(row=i, column=3).value
        visitante = sh.cell(row=i, column=4).value
        placar_mandante = int(sh.cell(row=i, column=5).value)
        placar_visitante = int(sh.cell(row=i, column=6).value)
        if placar_mandante == placar_visitante:
            total_pontos_mandante = total_pontos_visitante = 1
        elif placar_mandante > placar_visitante:
            total_pontos_mandante = 3
            total_pontos_visitante = 0
        else:
            total_pontos_mandante = 0
            total_pontos_visitante = 3
        classificacao_mandante = {'time': mandante, 'total_pontos': total_pontos_mandante, 'gols_pro': placar_mandante, 'gols_contra': placar_visitante, 'saldo_gols': placar_mandante-placar_visitante}
        classificacao_visitante = {'time': visitante, 'total_pontos': total_pontos_visitante}
        print(cell_obj.value, end=" ")